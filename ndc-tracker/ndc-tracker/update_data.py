#!/usr/bin/env python3
"""
NDC Transport Tracker - Data Processing Script
Processes the GIZ-SLOCAT Excel database and generates JSON for the dashboard
"""

import openpyxl
import json
from collections import defaultdict, Counter
from pathlib import Path

def process_excel(excel_path):
    """Process Excel file and return structured data"""
    
    print("📊 Opening Excel file...")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    document_sheet = wb['Document']
    country_sheet = wb['Country']
    mitigation_sheet = wb['Mitigation']
    
    # Column indices
    col_country = 3
    col_type = 5
    col_version = 8
    col_transport = 20
    col_status = 10
    col_region = 26
    col_doc_id = 1
    
    print("🔄 Processing Tab 1 data (NDC Progress)...")
    
    # ========================================================================
    # TAB 1: Progress in NDC Transport Targets
    # ========================================================================
    
    data_by_generation = {}
    countries_data = {}
    
    for gen_name, gen_versions in [
        ('gen1', ['1.0', '1.1']),
        ('gen2', ['2.0', '2.1']),
        ('gen3', ['3.0'])
    ]:
        total_ndcs = set()
        ndcs_with_transport = set()
        regions_total = defaultdict(set)
        regions_transport = defaultdict(set)
        
        for row in range(2, document_sheet.max_row + 1):
            country = document_sheet.cell(row, col_country).value
            doc_type = document_sheet.cell(row, col_type).value
            version = document_sheet.cell(row, col_version).value
            has_transport = document_sheet.cell(row, col_transport).value
            status = document_sheet.cell(row, col_status).value
            region = document_sheet.cell(row, col_region).value
            
            if doc_type != 'NDC' or not country or not version:
                continue
            
            # Exclude countries covered by EU
            if status and 'covered by eu' in str(status).lower():
                continue
            
            # Check if this version matches current generation
            is_gen = any(v in str(version) for v in gen_versions)
            if not is_gen:
                continue
            
            total_ndcs.add(country)
            
            if region:
                regions_total[str(region)].add(country)
            
            # Store country data for map
            if country not in countries_data:
                countries_data[country] = {
                    'name': country,
                    'region': str(region) if region else 'Unknown',
                    'generations': {}
                }
            
            countries_data[country]['generations'][gen_name] = {
                'has_transport': str(has_transport).lower() == 'yes',
                'version': str(version)
            }
            
            if has_transport and str(has_transport).lower() == 'yes':
                ndcs_with_transport.add(country)
                if region:
                    regions_transport[str(region)].add(country)
        
        data_by_generation[gen_name] = {
            'total_submitted': len(total_ndcs),
            'with_transport': len(ndcs_with_transport),
            'regions': {}
        }
        
        # Regional breakdown
        for region in regions_total:
            data_by_generation[gen_name]['regions'][region] = {
                'total': len(regions_total[region]),
                'with_transport': len(regions_transport.get(region, set()))
            }
    
    print(f"   ✓ Processed {len(countries_data)} countries")
    
    print("🔄 Processing Tab 2 data (Leading Measures)...")
    
    # ========================================================================
    # TAB 2: Leading Measures for Decarbonisation
    # ========================================================================
    
    # Identify latest active NDCs
    latest_active_ndcs = {}
    
    for row in range(2, document_sheet.max_row + 1):
        country = document_sheet.cell(row, 3).value
        doc_type = document_sheet.cell(row, 5).value
        version = document_sheet.cell(row, 8).value
        status = document_sheet.cell(row, 10).value
        doc_id = document_sheet.cell(row, 1).value
        
        if doc_type != 'NDC' or not country or not version:
            continue
        
        if status and 'covered by eu' in str(status).lower():
            continue
        
        if status != 'Active':
            continue
        
        # Priority: 3.0 > 2.0 > 1.0
        current_priority = 0
        if country in latest_active_ndcs:
            if '3.0' in latest_active_ndcs[country]['version']:
                current_priority = 3
            elif '2.0' in latest_active_ndcs[country]['version']:
                current_priority = 2
            else:
                current_priority = 1
        
        new_priority = 0
        if '3.0' in str(version):
            new_priority = 3
        elif '2.0' in str(version):
            new_priority = 2
        else:
            new_priority = 1
        
        if new_priority >= current_priority:
            latest_active_ndcs[country] = {
                'doc_id': doc_id,
                'version': str(version)
            }
    
    latest_doc_ids = {v['doc_id'] for v in latest_active_ndcs.values()}
    
    # Calculate measures by category (Latest Active NDCs)
    measures_by_category = defaultdict(lambda: {'ndcs': set(), 'countries': set(), 'mentions': 0})
    measures_by_country = defaultdict(lambda: {'total': 0, 'categories': Counter()})
    
    # Also calculate by generation
    measures_by_gen = {
        'gen1': defaultdict(lambda: {'ndcs': set(), 'mentions': 0}),
        'gen2': defaultdict(lambda: {'ndcs': set(), 'mentions': 0}),
        'gen3': defaultdict(lambda: {'ndcs': set(), 'mentions': 0}),
        'all': defaultdict(lambda: {'ndcs': set(), 'mentions': 0})
    }
    
    col_category = 10
    col_version_mit = 7
    
    for row in range(2, mitigation_sheet.max_row + 1):
        doc_id = mitigation_sheet.cell(row, 1).value
        country = mitigation_sheet.cell(row, 3).value
        category = mitigation_sheet.cell(row, col_category).value
        version = mitigation_sheet.cell(row, col_version_mit).value
        
        if not category or not country:
            continue
        
        cat_str = str(category)
        
        # For latest active
        if doc_id in latest_doc_ids:
            ndc_id = f"{country}_{doc_id}"
            measures_by_category[cat_str]['ndcs'].add(ndc_id)
            measures_by_category[cat_str]['countries'].add(country)
            measures_by_category[cat_str]['mentions'] += 1
            
            measures_by_country[country]['total'] += 1
            measures_by_country[country]['categories'][cat_str] += 1
        
        # By generation
        if version:
            gen_key = None
            if '1.0' in str(version) or '1.1' in str(version):
                gen_key = 'gen1'
            elif '2.0' in str(version) or '2.1' in str(version):
                gen_key = 'gen2'
            elif '3.0' in str(version):
                gen_key = 'gen3'
            
            if gen_key:
                ndc_id = f"{country}_{doc_id}"
                measures_by_gen[gen_key][cat_str]['ndcs'].add(ndc_id)
                measures_by_gen[gen_key][cat_str]['mentions'] += 1
                measures_by_gen['all'][cat_str]['ndcs'].add(ndc_id)
                measures_by_gen['all'][cat_str]['mentions'] += 1
    
    wb.close()
    
    print(f"   ✓ Processed {len(measures_by_category)} measure categories")
    
    # ========================================================================
    # Build final JSON structure
    # ========================================================================
    
    output_data = {
        'metadata': {
            'total_possible_ndcs': 169,
            'last_updated': 'auto-generated',
            'data_source': 'GIZ-SLOCAT Transport Tracker Database'
        },
        
        'tab1': {
            'generations': {
                'gen1': {
                    'name': 'First Generation',
                    'period': '2015-2019',
                    **data_by_generation['gen1']
                },
                'gen2': {
                    'name': 'Second Generation',
                    'period': '2020-2024',
                    **data_by_generation['gen2']
                },
                'gen3': {
                    'name': 'Third Generation',
                    'period': '2024-ongoing',
                    **data_by_generation['gen3']
                }
            },
            'countries': {}
        },
        
        'tab2': {
            'categories_latest': {},
            'categories_by_generation': {},
            'countries': {}
        }
    }
    
    # Convert countries_data to serializable format
    for country, data in countries_data.items():
        output_data['tab1']['countries'][country] = {
            'name': country,
            'region': data['region'],
            'generations': data['generations']
        }
    
    # Categories (latest active)
    for cat, data in measures_by_category.items():
        output_data['tab2']['categories_latest'][cat] = {
            'ndcs_count': len(data['ndcs']),
            'countries_count': len(data['countries']),
            'mentions': data['mentions']
        }
    
    # Categories by generation
    for gen_key, categories in measures_by_gen.items():
        output_data['tab2']['categories_by_generation'][gen_key] = {}
        for cat, data in categories.items():
            output_data['tab2']['categories_by_generation'][gen_key][cat] = {
                'ndcs_count': len(data['ndcs']),
                'mentions': data['mentions']
            }
    
    # Countries (for heat map tab 2)
    for country, data in measures_by_country.items():
        top_categories = data['categories'].most_common(3)
        output_data['tab2']['countries'][country] = {
            'total_mentions': data['total'],
            'top_categories': [{'category': cat, 'count': count} for cat, count in top_categories]
        }
    
    return output_data


def main():
    """Main execution function"""
    
    print("=" * 80)
    print("🚀 NDC Transport Tracker - Data Update Script")
    print("=" * 80)
    
    # Find Excel file
    data_dir = Path('data')
    excel_files = list(data_dir.glob('*.xlsx'))
    
    if not excel_files:
        print("❌ Error: No Excel file found in data/ directory")
        print("   Please place your Excel file in the data/ folder")
        return 1
    
    excel_path = excel_files[0]
    print(f"\n📂 Found Excel file: {excel_path.name}")
    
    # Process data
    try:
        data = process_excel(excel_path)
    except Exception as e:
        print(f"\n❌ Error processing Excel: {e}")
        return 1
    
    # Save JSON
    output_dir = data_dir / 'processed'
    output_dir.mkdir(exist_ok=True)
    output_path = output_dir / 'data.json'
    
    print(f"\n💾 Saving JSON to {output_path}...")
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    
    # Summary
    print("\n" + "=" * 80)
    print("✅ DATA UPDATE COMPLETE!")
    print("=" * 80)
    print(f"   • Countries processed: {len(data['tab1']['countries'])}")
    print(f"   • Measure categories: {len(data['tab2']['categories_latest'])}")
    print(f"   • Output file: {output_path}")
    print(f"   • File size: {output_path.stat().st_size:,} bytes")
    print("\n🎉 Dashboard data is ready!")
    
    return 0


if __name__ == '__main__':
    exit(main())
