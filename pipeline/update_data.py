#!/usr/bin/env python3
"""
NDC Transport Tracker - Data Processing Script
Processes the GIZ-SLOCAT Excel database and generates:
  - data/processed/data.json (for main tracker dashboard)
  - data/processed/comparison-data.json (for comparison dashboard)
"""

import json
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path

import openpyxl


# ============================================================================
# Helpers
# ============================================================================

def get_gen(version):
    """Map a version string like 'NDC 2.1' to 'gen1' / 'gen2' / 'gen3'."""
    v = str(version).strip()
    if v.startswith("NDC 1"): return "gen1"
    if v.startswith("NDC 2"): return "gen2"
    if v.startswith("NDC 3"): return "gen3"
    return None


# ============================================================================
# Schema validation — fail fast with clear errors instead of mid-parse crashes
# ============================================================================

REQUIRED_SCHEMA = {
    # sheet name: columns the pipeline reads (checked before any parsing)
    "Country":    ["Country Codes", "Country", "Region", "Income Level Group"],
    "Document":   ["Country Code", "Document ID", "Type of document",
                   "Document name", "Version number", "Date", "Status", "URL",
                   "Transport content", "Contains transport mitigation target",
                   "Contains transport adaptation target",
                   "Contains transport mitigation measures",
                   "Contains transport adaptation measures"],
    "Targets":    ["Country Code", "Document ID", "Target area", "Target scope",
                   "Target type", "Target Year", "GHG target?",
                   "Conditionality", "Content"],
    "Mitigation": ["Country Code", "Document ID", "Category", "Purpose",
                   "Instrument", "A-S-I", "Quote"],
    "Adaptation": ["Country Code", "Document ID"],
    "Benefits":   ["Country Code", "Document ID"],
    "References": ["Country Code", "Document ID"],
}


def validate_workbook(wb):
    """Check required sheets and columns exist BEFORE parsing.
    Reports every problem at once, then aborts — so one failed run reveals
    all schema issues instead of one cryptic KeyError at a time."""
    problems = []
    for sheet, cols in REQUIRED_SCHEMA.items():
        if sheet not in wb.sheetnames:
            problems.append(f"Sheet '{sheet}' is missing from the workbook")
            continue
        header = next(wb[sheet].iter_rows(values_only=True))
        headers = {str(h).strip() for h in header if h is not None}
        for col in cols:
            if col not in headers:
                problems.append(f"Sheet '{sheet}' is missing column '{col}'")
    if problems:
        print("❌  DATABASE SCHEMA CHECK FAILED — the Excel structure changed:")
        for p in problems:
            print(f"    • {p}")
        print("    Fix the Excel (or update REQUIRED_SCHEMA in this script if")
        print("    the change was intentional) and push again.")
        sys.exit(1)
    print("✅  Schema check passed — all required sheets and columns present\n")


# ============================================================================
# Main tracker processing (EXISTING - UNCHANGED)
# ============================================================================

def process_excel(excel_path):
    print("📊  Opening Excel file…")
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)

    doc_sheet     = wb["Document"]
    mit_sheet     = wb["Mitigation"]
    targets_sheet = wb["Targets"]

    EU_STATUSES = {"Covered by EU", "Covered by EU archived"}

    # Column positions (0-indexed)
    D_DOCID     = 0
    D_CODE      = 1
    D_NAME      = 2
    D_TYPE      = 4
    D_VERSION   = 7
    D_STATUS    = 9
    D_TRANSPORT = 10
    D_REGION    = 25

    M_DOCID    = 0
    M_CODE     = 1
    M_VERSION  = 6
    M_CATEGORY = 9

    # Targets sheet columns
    T_DOCID       = 0
    T_CODE        = 1
    T_TYPE        = 3
    T_TARGET_AREA = 8
    T_GHG_FLAG    = 10

    # ── Pass 1: Document sheet ─────────────────────────────────────────
    country_master = {}
    doc_id_info    = {}

    for row in doc_sheet.iter_rows(min_row=2, values_only=True):
        if not row[D_DOCID]: break

        doc_id  = row[D_DOCID]
        code    = row[D_CODE]
        name    = row[D_NAME]
        dtype   = row[D_TYPE]
        version = row[D_VERSION]
        status  = row[D_STATUS]
        has_t   = row[D_TRANSPORT]
        region  = row[D_REGION]

        if dtype != "NDC" or not code or not version: continue
        if status in EU_STATUSES: continue

        gen = get_gen(version)
        if not gen: continue

        code   = str(code).strip()
        status = str(status).strip()

        if code not in country_master:
            country_master[code] = {
                "name":   str(name).strip() if name else code,
                "region": str(region).strip() if region else "Unknown",
            }

        doc_id_info[doc_id] = {
            "code":      code,
            "gen":       gen,
            "status":    status,
            "has_t":     str(has_t).strip().lower() == "yes",
            "region":    str(region).strip() if region else "Unknown",
        }

    print(f"   ✓ {len(country_master)} countries, {len(doc_id_info)} NDC documents")

    # ── Latest active doc per country ──────────────────────────────────
    GEN_PRIORITY = {"gen1": 1, "gen2": 2, "gen3": 3}
    country_best = {}

    for doc_id, info in doc_id_info.items():
        if info["status"] != "Active": continue
        code = info["code"]
        prio = GEN_PRIORITY[info["gen"]]
        if code not in country_best or prio > country_best[code][0]:
            country_best[code] = (prio, doc_id)

    latest_doc_ids = {v[1] for v in country_best.values()}
    latest_gen_map = {
        code: ["gen1", "gen2", "gen3"][prio - 1]
        for code, (prio, _) in country_best.items()
    }

    # ── Per-gen active/archived country counts ─────────────────────────
    gen_active_codes   = {g: set() for g in ["gen1", "gen2", "gen3"]}
    gen_archived_codes = {g: set() for g in ["gen1", "gen2", "gen3"]}

    for info in doc_id_info.values():
        code = info["code"]; gen = info["gen"]
        if info["status"] == "Active":
            gen_active_codes[gen].add(code)
        else:
            gen_archived_codes[gen].add(code)

    gen_counts = {
        g: {
            "active":   len(gen_active_codes[g]),
            "archived": len(gen_archived_codes[g]),
        }
        for g in ["gen1", "gen2", "gen3"]
    }
    gen_counts["latest"] = {"active": len(latest_doc_ids), "archived": 0}

    # ── Targets sheet: identify transport and GHG transport doc_ids ────
    print("🔄  Scanning Targets sheet…")

    doc_has_transport     = set()  # any transport sector mitigation target
    doc_has_ghg_transport = set()  # quantified GHG transport target

    for row in targets_sheet.iter_rows(min_row=2, values_only=True):
        if not row[T_DOCID]: break

        doc_id      = row[T_DOCID]
        doc_type    = row[T_TYPE]
        target_area = str(row[T_TARGET_AREA] or "")
        ghg_flag    = str(row[T_GHG_FLAG] or "")

        if doc_type != "NDC": continue
        if doc_id not in doc_id_info: continue

        if "Transport sector mitigation" in target_area:
            doc_has_transport.add(doc_id)
            if ghg_flag == "GHG":
                doc_has_ghg_transport.add(doc_id)

    print(f"   ✓ {len(doc_has_transport)} docs with transport target, "
          f"{len(doc_has_ghg_transport)} with GHG transport target")

    # ── TAB 1: per-generation stats + per-country data ─────────────────
    print("🔄  Computing Tab 1 data…")

    # Per gen: collect unique countries and their transport flags
    country_gen_docs = defaultdict(lambda: defaultdict(list))
    for doc_id, info in doc_id_info.items():
        country_gen_docs[info["code"]][info["gen"]].append(doc_id)

    GEN_META = {
        "gen1": {"name": "First Generation",  "period": "2015–2019"},
        "gen2": {"name": "Second Generation", "period": "2020–2024"},
        "gen3": {"name": "Third Generation",  "period": "2024–ongoing"},
    }

    tab1_generations = {}
    for gen in ["gen1", "gen2", "gen3"]:
        submitted        = set()
        with_transport   = set()
        with_ghg         = set()
        reg_submitted    = defaultdict(set)
        reg_transport    = defaultdict(set)
        reg_ghg          = defaultdict(set)

        for code, gen_docs in country_gen_docs.items():
            if gen not in gen_docs: continue
            docs   = gen_docs[gen]
            region = doc_id_info[docs[0]]["region"]
            submitted.add(code)
            reg_submitted[region].add(code)

            for doc_id in docs:
                if doc_id in doc_has_transport:
                    with_transport.add(code)
                    reg_transport[region].add(code)
                if doc_id in doc_has_ghg_transport:
                    with_ghg.add(code)
                    reg_ghg[region].add(code)

        regions_out = {
            reg: {
                "total":              len(reg_submitted[reg]),
                "with_transport":     len(reg_transport.get(reg, set())),
                "with_ghg_transport": len(reg_ghg.get(reg, set())),
            }
            for reg in reg_submitted
        }

        tab1_generations[gen] = {
            **GEN_META[gen],
            "total_submitted":     len(submitted),
            "with_transport":      len(with_transport),
            "with_ghg_transport":  len(with_ghg),
            "regions":             regions_out,
        }

    # Per-country: latest_has_transport, latest_has_ghg_transport
    countries_tab1 = {}

    for row in doc_sheet.iter_rows(min_row=2, values_only=True):
        if not row[D_DOCID]: break

        code    = row[D_CODE]
        dtype   = row[D_TYPE]
        version = row[D_VERSION]
        status  = row[D_STATUS]
        has_t   = row[D_TRANSPORT]
        region  = row[D_REGION]

        if dtype != "NDC" or not code or not version: continue
        if status in EU_STATUSES: continue

        gen = get_gen(version)
        if not gen: continue

        code    = str(code).strip()
        has_t_b = str(has_t or "").strip().lower() == "yes"
        region  = str(region or "Unknown").strip()

        if code not in countries_tab1:
            countries_tab1[code] = {
                "iso3":                    code,
                "name":                    country_master[code]["name"],
                "region":                  country_master[code]["region"],
                "latest_active_gen":       latest_gen_map.get(code),
                "latest_has_transport":    None,
                "latest_has_ghg_transport": False,
                "generations":             {},
            }

        gd = countries_tab1[code]["generations"]
        if gen not in gd:
            gd[gen] = {"has_transport": has_t_b, "version": str(version).strip()}
        elif has_t_b:
            gd[gen]["has_transport"] = True

    # Fill latest transport flags from latest active doc
    for code, (prio, doc_id) in country_best.items():
        if code not in countries_tab1: continue
        lat_gen = latest_gen_map.get(code)
        cd = countries_tab1[code]
        if lat_gen and lat_gen in cd["generations"]:
            cd["latest_has_transport"] = cd["generations"][lat_gen]["has_transport"]
        cd["latest_has_ghg_transport"] = doc_id in doc_has_ghg_transport

    # ── Calculate had_transport_previously flag ───────────────────────────
    print("🔄  Computing 'had_transport_previously' flags…")

    for code, cd in countries_tab1.items():
        latest_gen = cd.get("latest_active_gen")
        latest_has_t = cd.get("latest_has_transport", False)
        
        # If latest already has transport, no need to check previous
        if latest_has_t:
            cd["had_transport_previously"] = False
            continue
        
        # Check if any earlier generation had transport
        had_previous = False
        for gen in ["gen1", "gen2", "gen3"]:
            # Skip if this is the latest generation or later than latest
            if not latest_gen:
                continue
            gen_priority = {"gen1": 1, "gen2": 2, "gen3": 3}
            if gen_priority.get(gen, 0) >= gen_priority.get(latest_gen, 0):
                continue
            
            # Check if this earlier generation had transport
            if cd["generations"].get(gen, {}).get("has_transport", False):
                had_previous = True
                break
        
        cd["had_transport_previously"] = had_previous

    withdrawn_count = sum(1 for cd in countries_tab1.values() if cd.get("had_transport_previously"))
    print(f"   ✓ {withdrawn_count} countries withdrew transport targets in latest NDC")

    print(f"   ✓ Tab1 built for {len(countries_tab1)} countries")

    # ── TAB 2: mitigation measures ─────────────────────────────────────
    print("🔄  Computing Tab 2 data…")

    cat_lgb_sets = defaultdict(lambda: {
        "gen1": set(), "gen2": set(), "gen3": set(), "mentions": 0
    })

    cat_gen_aa_sets = {
        g: defaultdict(lambda: {
            "active_c": set(), "archived_c": set(),
            "active_m": 0,     "archived_m": 0,
        })
        for g in ["gen1", "gen2", "gen3"]
    }

    country_gen_cats    = {}
    country_latest_cats = {}
    country_gen_asi     = {}   # NEW: A-S-I counts per country per generation
    country_latest_asi  = {}   # NEW: A-S-I counts per country, latest NDC

    M_ASI = 13  # 'A-S-I' column in the Mitigation sheet

    def _asi_parts(raw):
        order = {"Avoid": 0, "Shift": 1, "Improve": 2}
        return [p.strip().capitalize() for p in str(raw or "").split(",")
                if p.strip().capitalize() in order]

    for row in mit_sheet.iter_rows(min_row=2, values_only=True):
        if not row[M_DOCID]: break

        doc_id   = row[M_DOCID]
        code     = row[M_CODE]
        version  = row[M_VERSION]
        category = row[M_CATEGORY]
        asi_raw  = row[M_ASI] if len(row) > M_ASI else None

        if not category or not code or not version or not doc_id: continue
        if doc_id not in doc_id_info: continue

        code = str(code).strip()
        cat  = str(category).strip()
        gen  = get_gen(version)
        if not gen: continue

        info   = doc_id_info[doc_id]
        status = info["status"]

        if doc_id in latest_doc_ids:
            lat_gen = latest_gen_map.get(code)
            if lat_gen:
                cat_lgb_sets[cat][lat_gen].add(code)
                cat_lgb_sets[cat]["mentions"] += 1
            if code not in country_latest_cats:
                country_latest_cats[code] = {}
            country_latest_cats[code][cat] = country_latest_cats[code].get(cat, 0) + 1
            for part in _asi_parts(asi_raw):
                country_latest_asi.setdefault(code, {})
                country_latest_asi[code][part] = \
                    country_latest_asi[code].get(part, 0) + 1

        for part in _asi_parts(asi_raw):
            country_gen_asi.setdefault(code, {}).setdefault(gen, {})
            country_gen_asi[code][gen][part] = \
                country_gen_asi[code][gen].get(part, 0) + 1

        s = cat_gen_aa_sets[gen][cat]
        if status == "Active":
            s["active_c"].add(code); s["active_m"] += 1
        else:
            s["archived_c"].add(code); s["archived_m"] += 1

        if code not in country_gen_cats:
            country_gen_cats[code] = {}
        if gen not in country_gen_cats[code]:
            country_gen_cats[code][gen] = {}
        country_gen_cats[code][gen][cat] = \
            country_gen_cats[code][gen].get(cat, 0) + 1

    cat_lgb_out = {
        cat: {
            "gen1_count": len(v["gen1"]),
            "gen2_count": len(v["gen2"]),
            "gen3_count": len(v["gen3"]),
            "mentions":   v["mentions"],
        }
        for cat, v in cat_lgb_sets.items()
    }

    cat_gen_aa_out = {
        gen: {
            cat: {
                "active_countries":   len(s["active_c"]),
                "archived_countries": len(s["archived_c"]),
                "active_mentions":    s["active_m"],
                "archived_mentions":  s["archived_m"],
            }
            for cat, s in cats.items()
        }
        for gen, cats in cat_gen_aa_sets.items()
    }

    categories_latest = {
        cat: {
            "countries_count": len(v["gen1"] | v["gen2"] | v["gen3"]),
            "mentions":        v["mentions"],
        }
        for cat, v in cat_lgb_sets.items()
    }

    print(f"   ✓ Tab2 built: {len(cat_lgb_out)} categories")

    # ── EU member states ───────────────────────────────────────────────
    print("🔄  Adding EU member states…")

    EU_CODE = "EEU"
    EU_MEMBER_ISO3 = [
        "AUT","BEL","BGR","HRV","CYP","CZE","DNK","EST","FIN","FRA",
        "DEU","GRC","HUN","IRL","ITA","LVA","LTU","LUX","MLT","NLD",
        "POL","PRT","ROU","SVK","SVN","ESP","SWE"
    ]
    EU_NAMES = {
        "AUT":"Austria","BEL":"Belgium","BGR":"Bulgaria","HRV":"Croatia",
        "CYP":"Cyprus","CZE":"Czechia","DNK":"Denmark","EST":"Estonia",
        "FIN":"Finland","FRA":"France","DEU":"Germany","GRC":"Greece",
        "HUN":"Hungary","IRL":"Ireland","ITA":"Italy","LVA":"Latvia",
        "LTU":"Lithuania","LUX":"Luxembourg","MLT":"Malta","NLD":"Netherlands",
        "POL":"Poland","PRT":"Portugal","ROU":"Romania","SVK":"Slovakia",
        "SVN":"Slovenia","ESP":"Spain","SWE":"Sweden"
    }

    eu_doc_ids_local = {
        doc_id: info for doc_id, info in doc_id_info.items()
        if info["code"] == EU_CODE
    }

    eu_best = None; eu_best_p = 0
    for doc_id, info in eu_doc_ids_local.items():
        if info["status"] != "Active": continue
        p = GEN_PRIORITY[info["gen"]]
        if p > eu_best_p:
            eu_best_p = p; eu_best = (doc_id, info)

    if eu_best:
        latest_eu_doc_id = eu_best[0]
        latest_eu_gen    = eu_best[1]["gen"]
        latest_eu_has_t  = eu_best[1].get("has_t", False)
        latest_eu_has_ghg = latest_eu_doc_id in doc_has_ghg_transport

        eu_gen_cats_local    = {}
        eu_latest_cats_local = {}

        for row in mit_sheet.iter_rows(min_row=2, values_only=True):
            if not row[M_DOCID]: break
            doc_id  = row[M_DOCID]
            code    = str(row[M_CODE] or "").strip()
            version = row[M_VERSION]
            cat     = row[M_CATEGORY]
            if code != EU_CODE or not cat: continue
            if doc_id not in eu_doc_ids_local: continue
            cat = str(cat).strip()
            gen = get_gen(version)
            if not gen: continue
            if gen not in eu_gen_cats_local:
                eu_gen_cats_local[gen] = {}
            eu_gen_cats_local[gen][cat] = eu_gen_cats_local[gen].get(cat, 0) + 1
            if doc_id == latest_eu_doc_id:
                eu_latest_cats_local[cat] = eu_latest_cats_local.get(cat, 0) + 1

        eu_gens_info = {}
        for row in doc_sheet.iter_rows(min_row=2, values_only=True):
            if not row[D_DOCID]: break
            if row[D_CODE] != EU_CODE or row[D_TYPE] != "NDC": continue
            gen = get_gen(row[D_VERSION])
            if not gen: continue
            has_t = str(row[D_TRANSPORT] or "").strip().lower() == "yes"
            ver   = str(row[D_VERSION]).strip()
            if gen not in eu_gens_info:
                eu_gens_info[gen] = {"has_transport": has_t, "version": ver}
            elif has_t:
                eu_gens_info[gen]["has_transport"] = True
                eu_gens_info[gen]["version"] = ver

        # Calculate had_transport_previously for EU
        eu_had_transport_previously = False
        if not latest_eu_has_t:
            for gen in ["gen1", "gen2", "gen3"]:
                gen_priority = {"gen1": 1, "gen2": 2, "gen3": 3}
                if gen_priority.get(gen, 0) >= gen_priority.get(latest_eu_gen, 0):
                    continue
                if eu_gens_info.get(gen, {}).get("has_transport", False):
                    eu_had_transport_previously = True
                    break

        for iso3 in EU_MEMBER_ISO3:
            countries_tab1[iso3] = {
                "iso3":                    iso3,
                "name":                    EU_NAMES.get(iso3, iso3),
                "region":                  "Europe",
                "latest_active_gen":       latest_eu_gen,
                "latest_has_transport":    latest_eu_has_t,
                "latest_has_ghg_transport": latest_eu_has_ghg,
                "had_transport_previously": eu_had_transport_previously,
                "generations":             eu_gens_info,
                "covered_by_eu":           True,
            }
            country_gen_cats[iso3]    = eu_gen_cats_local.copy()
            country_latest_cats[iso3] = eu_latest_cats_local.copy()
            latest_gen_map[iso3]      = latest_eu_gen

        print(f"   ✓ Added {len(EU_MEMBER_ISO3)} EU member states "
              f"(latest: {latest_eu_gen}, transport: {latest_eu_has_t}, GHG: {latest_eu_has_ghg}, "
              f"had_previous: {eu_had_transport_previously})")

    # ── Assemble output ────────────────────────────────────────────────
    output = {
        "metadata": {
            "total_possible_ndcs":    169,
            "last_updated":           str(date.today()),
            "data_source":            "GIZ-SLOCAT Transport Tracker Database",
            "gen_counts":             gen_counts,
            "region_possible_ndcs":   {
                reg: sum(
                    1 for cd in countries_tab1.values()
                    if cd.get("region") == reg and not cd.get("covered_by_eu")
                )
                for reg in ["Africa","Asia","Europe",
                            "Latin America and the Caribbean",
                            "Northern America","Oceania"]
            },
        },
        "tab1": {
            "generations": tab1_generations,
            "countries":   countries_tab1,
        },
        "tab2": {
            "categories_latest":        categories_latest,
            "cat_latest_gen_breakdown": cat_lgb_out,
            "cat_gen_active_archived":  cat_gen_aa_out,
            "country_gen_cats":         country_gen_cats,
            "country_latest_cats":      country_latest_cats,
            "country_gen_asi":          country_gen_asi,
            "country_latest_asi":       country_latest_asi,
            "latest_gen_map":           latest_gen_map,
        },
    }

    # ── Per-country transport emissions (Mt CO₂e, EDGAR 2023) for KPIs ──
    try:
        ctry_sheet = wb["Country"]
        for row in ctry_sheet.iter_rows(min_row=2, values_only=True):
            code = str(row[1] or "").strip()
            if code in countries_tab1 and isinstance(row[28], (int, float)):
                countries_tab1[code]["ghg_transport"] = round(row[28], 2)
    except Exception as exc:
        print(f"   ⚠ Could not attach per-country emissions: {exc}")

    wb.close()
    return output


# ============================================================================
# Comparison dashboard processing (NEW)
# ============================================================================

def process_comparison_data(excel_path):
    """
    Process Excel data for the comparison dashboard.
    Extracts NDC documents grouped by country and generation with all relevant details.
    """
    print("🔄  Processing comparison data…")
    
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    
    doc_sheet = wb["Document"]
    targets_sheet = wb["Targets"]
    mitigation_sheet = wb["Mitigation"]
    adaptation_sheet = wb["Adaptation"]
    
    # Get all mode column headers from Mitigation sheet (columns P to AL)
    mit_headers = [cell.value for cell in mitigation_sheet[1]]
    mit_mode_cols = {}
    for idx, header in enumerate(mit_headers):
        if idx >= 15 and idx <= 37:  # P to AL (0-indexed: 15-37)
            if header:
                mit_mode_cols[idx] = str(header).strip()
    
    # Get all mode column headers from Adaptation sheet (columns M to AI)
    adapt_headers = [cell.value for cell in adaptation_sheet[1]]
    adapt_mode_cols = {}
    for idx, header in enumerate(adapt_headers):
        if idx >= 12 and idx <= 34:  # M to AI (0-indexed: 12-34)
            if header:
                adapt_mode_cols[idx] = str(header).strip()
    
    # Column indices (0-indexed)
    D_DOCID = 0
    D_CODE = 1
    D_NAME = 2
    D_TYPE = 4
    D_VERSION = 7
    D_DATE = 8
    D_STATUS = 9
    D_TRANSPORT = 10
    D_NETZERO = 11
    D_URL = 23  # Column X
    
    T_DOCID = 0
    T_AREA = 8
    T_SCOPE = 9     # Column J — "Target scope" (not previously captured)
    T_GHG = 10      # Column K (was incorrectly 9)
    T_TYPE = 11     # Column L (was incorrectly 10)
    T_COND = 12     # Column M (was incorrectly 11)
    T_YEAR = 13     # Column N (was incorrectly 12)
    T_CONTENT = 14
    T_PAGE = 15     # Column P
    
    M_DOCID = 0
    M_CAT = 9
    M_PURPOSE = 10    # Column K — "Purpose" (not previously captured)
    M_INSTRUMENT = 11 # Column L — "Instrument" (not previously captured)
    M_QUOTE = 12
    M_ASI = 13
    M_PAGE = 14     # Column O
    
    A_DOCID = 0
    A_CAT = 8
    A_MEASURE = 9    # Column J — "Measure" (not previously captured)
    A_QUOTE = 10
    A_PAGE = 11     # Column L
    
    # Process documents
    documents = {}
    
    for row in doc_sheet.iter_rows(min_row=2, values_only=True):
        if not row[D_DOCID]:
            break
        
        doc_id = row[D_DOCID]
        code = row[D_CODE]
        name = row[D_NAME]
        dtype = row[D_TYPE]
        version = row[D_VERSION]
        date_val = row[D_DATE]
        status = row[D_STATUS]
        has_transport = row[D_TRANSPORT]
        has_netzero = row[D_NETZERO]
        url = row[D_URL]
        
        if dtype != "NDC" or not code or not version:
            continue
        
        code = str(code).strip()
        version_str = str(version).strip()
        
        # Determine generation
        gen = get_gen(version_str)
        if not gen:
            continue
        
        # Format date - convert Excel serial date to MM.YYYY
        date_str = ""
        if date_val:
            if hasattr(date_val, 'strftime'):
                # If it's already a datetime object
                date_str = date_val.strftime('%m.%Y')
            elif isinstance(date_val, (int, float)):
                # If it's an Excel serial date number
                from datetime import datetime, timedelta
                try:
                    excel_epoch = datetime(1899, 12, 30)
                    converted_date = excel_epoch + timedelta(days=float(date_val))
                    date_str = converted_date.strftime('%m.%Y')
                except:
                    date_str = str(date_val)
            else:
                date_str = str(date_val)
        
        documents[doc_id] = {
            "doc_id": doc_id,
            "country_code": code,
            "country_name": str(name).strip() if name else code,
            "doc_type": str(dtype).strip(),
            "version": version_str,
            "generation": gen,
            "date": date_str,
            "status": str(status).strip() if status else "",
            "has_transport_target": str(has_transport).strip().lower() == "yes",
            "has_netzero_target": str(has_netzero).strip().lower() == "yes",
            "url": str(url).strip() if url else "",
        }
    
    # Process targets
    targets_by_doc = {}
    
    for row in targets_sheet.iter_rows(min_row=2, values_only=True):
        if not row[T_DOCID]:
            break
        
        doc_id = row[T_DOCID]
        target_area = row[T_AREA]
        target_scope = row[T_SCOPE]
        ghg = row[T_GHG]
        target_type = row[T_TYPE]
        cond = row[T_COND]
        year = row[T_YEAR]
        content = row[T_CONTENT]
        page = row[T_PAGE]
        
        if doc_id not in documents:
            continue
        
        if not target_area:
            continue
        
        target_area_str = str(target_area).strip()
        
        # Only include these three target types
        if target_area_str not in [
            "Transport sector mitigation target",
            "Transport sector adaptation target",
            "Net zero target"
        ]:
            continue
        
        if doc_id not in targets_by_doc:
            targets_by_doc[doc_id] = []
        
        targets_by_doc[doc_id].append({
            "target_area": target_area_str,
            "target_scope": str(target_scope).strip() if target_scope else "—",
            "ghg_target": str(ghg).strip() if ghg else "—",
            "target_type": str(target_type).strip() if target_type else "—",
            "conditionality": str(cond).strip() if cond else "—",
            "target_year": str(year).strip() if year else "—",
            "content": str(content).strip() if content else "—",
            "page": str(page).strip() if page else "",
        })
    
    # Process mitigation measures
    mitigation_by_doc = {}
    
    for row in mitigation_sheet.iter_rows(min_row=2, values_only=True):
        if not row[M_DOCID]:
            break
        
        doc_id = row[M_DOCID]
        category = row[M_CAT]
        purpose = row[M_PURPOSE]
        instrument = row[M_INSTRUMENT]
        quote = row[M_QUOTE]
        asi = row[M_ASI]
        page = row[M_PAGE]
        
        if doc_id not in documents:
            continue
        
        if not category:
            continue
        
        category_str = str(category).strip()
        
        # Get modes marked with X
        modes = []
        for col_idx, mode_name in mit_mode_cols.items():
            if col_idx < len(row) and str(row[col_idx]).strip().upper() == "X":
                modes.append(mode_name)
        
        # Keep ASI value as-is (e.g., "Improve", "Avoid-Shift", etc.)
        asi_str = "—"
        if asi:
            asi_val = str(asi).strip()
            if asi_val and asi_val != "—":
                asi_str = asi_val
        
        if doc_id not in mitigation_by_doc:
            mitigation_by_doc[doc_id] = {}
        
        if category_str not in mitigation_by_doc[doc_id]:
            mitigation_by_doc[doc_id][category_str] = []
        
        mitigation_by_doc[doc_id][category_str].append({
            "quote": str(quote).strip() if quote else "—",
            "purpose": str(purpose).strip() if purpose else "—",
            "instrument": str(instrument).strip() if instrument else "—",
            "asi": asi_str,
            "modes": ", ".join(modes) if modes else "—",
            "page": str(page).strip() if page else "",
        })
    
    # Process adaptation measures
    adaptation_by_doc = {}
    
    for row in adaptation_sheet.iter_rows(min_row=2, values_only=True):
        if not row[A_DOCID]:
            break
        
        doc_id = row[A_DOCID]
        category = row[A_CAT]
        measure = row[A_MEASURE]
        quote = row[A_QUOTE]
        page = row[A_PAGE]
        
        if doc_id not in documents:
            continue
        
        if not category:
            continue
        
        category_str = str(category).strip()
        
        # Get modes marked with X
        modes = []
        for col_idx, mode_name in adapt_mode_cols.items():
            if col_idx < len(row) and str(row[col_idx]).strip().upper() == "X":
                modes.append(mode_name)
        
        if doc_id not in adaptation_by_doc:
            adaptation_by_doc[doc_id] = {}
        
        if category_str not in adaptation_by_doc[doc_id]:
            adaptation_by_doc[doc_id][category_str] = []
        
        adaptation_by_doc[doc_id][category_str].append({
            "quote": str(quote).strip() if quote else "—",
            "measure": str(measure).strip() if measure else "—",
            "modes": ", ".join(modes) if modes else "—",
            "page": str(page).strip() if page else "",
        })
    
    # Group by country
    countries_data = {}
    
    for doc_id, doc in documents.items():
        code = doc["country_code"]
        gen = doc["generation"]
        
        if code not in countries_data:
            countries_data[code] = {
                "country_name": doc["country_name"],
                "generations": {
                    "gen1": [],
                    "gen2": [],
                    "gen3": []
                }
            }
        
        # Add all data for this document
        doc_full = {
            **doc,
            "targets": targets_by_doc.get(doc_id, []),
            "mitigation_measures": mitigation_by_doc.get(doc_id, {}),
            "adaptation_measures": adaptation_by_doc.get(doc_id, {}),
        }
        
        countries_data[code]["generations"][gen].append(doc_full)
    
    # Sort versions within each generation (latest first)
    for country in countries_data.values():
        for gen in ["gen1", "gen2", "gen3"]:
            country["generations"][gen].sort(
                key=lambda x: x["version"],
                reverse=True
            )
    
    wb.close()
    
    print(f"   ✓ Processed {len(countries_data)} countries for comparison")
    
    return {
        "countries": countries_data,
        "last_updated": str(date.today()),
    }


# ============================================================================
# Entry point
# ============================================================================

def run_dashboards(excel_path):
    """Main tracker dashboard + comparison dashboard outputs."""
    data = process_excel(excel_path)
    comparison_data = process_comparison_data(excel_path)

    output_dir = Path("data") / "processed"
    output_dir.mkdir(parents=True, exist_ok=True)

    json_path = output_dir / "data.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print(f"💾  data.json saved ({json_path.stat().st_size:,} bytes)")

    comparison_path = output_dir / "comparison-data.json"
    with open(comparison_path, "w", encoding="utf-8") as f:
        json.dump(comparison_data, f, indent=2, ensure_ascii=False)
    print(f"💾  comparison-data.json saved ({comparison_path.stat().st_size:,} bytes)")

    print(f"   Countries: {len(data['tab1']['countries'])} · "
          f"Comparison: {len(comparison_data['countries'])}")
    return data, comparison_data




# ============================================================================
# COUNTRY PROFILES (profiles/)
# ============================================================================

import re
import unicodedata
from datetime import datetime, timedelta


try:
    import pycountry
except ImportError:
    pycountry = None

# ============================================================================
# Constants
# ============================================================================

EXCEL_DEFAULT = "GIZ-SLOCAT_Transport-Tracker-database.xlsx"
OUT_DIR = Path("profiles/data/countries")

EU_CODE = "EEU"
EU_STATUSES = {"Covered by EU", "Covered by EU archived"}
EU_MEMBER_ISO3 = [
    "AUT", "BEL", "BGR", "HRV", "CYP", "CZE", "DNK", "EST", "FIN", "FRA",
    "DEU", "GRC", "HUN", "IRL", "ITA", "LVA", "LTU", "LUX", "MLT", "NLD",
    "POL", "PRT", "ROU", "SVK", "SVN", "ESP", "SWE",
]

# ISO3 codes not resolvable via pycountry (synthetic / legacy codes in the DB)
ISO2_OVERRIDES = {
    "EEU": "eu",   # European Union — flagcdn supports "eu"
    "XKX": "xk",   # Kosovo
}

MODE_COLUMNS = [
    "Informal transport", "Active mobility", "Walking", "Cycling", "Road",
    "Two-/Three-wheelers", "Cars", "Private cars", "Taxis", "Truck", "Bus",
    "Rail", "Heavy rail", "High-speed rail", "Transit rail", "Water",
    "Coastal shipping", "Inland shipping", "International maritime",
    "Aviation", "Domestic aviation", "International aviation",
]
GEO_COLUMNS = ["Urban", "Rural", "Inter-city"]

COALITION_COLUMNS = {
    "Declaration on accelerating the transition to 100% zero emission cars and vans":
        "ZEV Declaration (100% zero-emission cars & vans)",
    "Breakthrough agenda - road transport": "Breakthrough Agenda — Road Transport",
    "International aviation climate ambition coalition":
        "International Aviation Climate Ambition Coalition",
    "Clydebank declaration for green shipping corridors":
        "Clydebank Declaration (green shipping corridors)",
    "Memorandum of understanding on zero-emission medium- and heavy-duty vehicles":
        "Global MoU on Zero-Emission MHDVs",
    "Beyond oil and gas alliance": "Beyond Oil & Gas Alliance",
    "Charge forward to zero emissions transportation":
        "Charge Forward to Zero Emissions Transportation",
    "Transport Decarbonisation Alliance": "Transport Decarbonisation Alliance",
    "International Zero-Emission Vehicle Alliance":
        "International Zero-Emission Vehicle Alliance",
    "ZEV Transition Council": "ZEV Transition Council",
    "Clean Energy Ministerial": "Clean Energy Ministerial",
}

EXCEL_EPOCH = date(1899, 12, 30)


# ============================================================================
# Helpers
# ============================================================================

def clean(v):
    """Normalise a cell value: strip strings, None for empty."""
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def excel_date(v):
    """Excel serial / datetime → ISO date string (or None)."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, (int, float)) and v > 20000:
        return (EXCEL_EPOCH + timedelta(days=int(v))).isoformat()
    return None


def yesno(v):
    """'yes'/'no'-style cells → True / False / None."""
    v = clean(v)
    if v is None:
        return None
    s = str(v).strip().lower()
    if s in ("yes", "y", "true", "x"):
        return True
    if s in ("no", "n", "false"):
        return False
    return None


def _profiles_get_gen(version):
    v = str(version or "").strip()
    if v.startswith("NDC 1"):
        return "gen1"
    if v.startswith("NDC 2"):
        return "gen2"
    if v.startswith("NDC 3"):
        return "gen3"
    return None


def norm_asi(v):
    """Normalise A-S-I labels: 'Avoid, shift, improve' → ['Avoid','Shift','Improve']."""
    v = clean(v)
    if not v:
        return []
    parts = [p.strip().capitalize() for p in str(v).split(",")]
    order = {"Avoid": 0, "Shift": 1, "Improve": 2}
    return sorted({p for p in parts if p in order}, key=lambda p: order[p])


def iso2_for(iso3):
    if iso3 in ISO2_OVERRIDES:
        return ISO2_OVERRIDES[iso3]
    if pycountry:
        c = pycountry.countries.get(alpha_3=iso3)
        if c:
            return c.alpha_2.lower()
    return None


def slugify(name):
    s = unicodedata.normalize("NFKD", str(name)).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9]+", "-", s.lower()).strip("-")


def rows_as_dicts(sheet):
    """Yield each data row as a dict keyed by (stripped) header names."""
    rows = sheet.iter_rows(values_only=True)
    header = next(rows)
    keys = [str(h).strip() if h is not None else f"col{i}" for i, h in enumerate(header)]
    for row in rows:
        if all(v is None for v in row):
            continue
        yield dict(zip(keys, row))


# ============================================================================
# GHG data loader — single source of truth: data/ghg.csv
# ============================================================================

def load_ghg_csv(path=None):
    """
    Single source of truth for GHG/transport emissions data.

    Reads data/ghg.csv (wide format: one row per country, columns repeating
    per year as "{Field}_{Year} [{unit}]") and returns everything
    build_profiles() and the trends chart need, from one parse:

        {
          "COL": {
            "snapshot": {
                "total_mt": 217.59,
                "transport_mt": 37.78,
                "transport_share_pct": 17.37,
                "transport_global_share_pct": 0.2138,
                "transport_per_capita": 0.733,
                "transport_sector_rank": 2,        # 1/2/3 if "Transport" is
                                                     # in Rank1/2/3_Sector for
                                                     # the latest year, else
                                                     # None (omitted on the
                                                     # profile page)
                "year": 2024,
                "source": "EDGAR",
            },
            "trends": {
                "years": [1970, 1971, ..., 2024],
                "total": [...],
                "transport": [...],
                "source": "EDGAR",
            },
          },
          ...
        }

    Replaces the prior two-file setup (data/ghg.json for the snapshot and
    data/sources/edgar_timeseries.csv for trends). Updating GHG data going
    forward only requires replacing data/ghg.csv and re-running the
    pipeline — no other file needs to change or stay in sync.

    The EU (EEU) snapshot/trends are summed from member states, matching
    the prior load_edgar_trends() behaviour (only years where at least 20
    of 27 member states reported are included).
    """
    import csv

    path = Path(path) if path else Path("data/ghg.csv")
    if not path.exists():
        print(f"   ⚠  {path} not found — emissions fall back to GIZ-SLOCAT Excel columns")
        return {}

    year_col_re = re.compile(r"^(.*?)_(\d{4})(?:\s*\[.*\])?$")

    def parse_float(raw):
        if raw is None:
            return None
        raw = raw.strip()
        if raw == "":
            return None
        try:
            return float(raw.replace(",", "."))
        except ValueError:
            return None

    per_country = {}  # iso3 -> {year: {field: value}}

    with path.open(encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            code = (row.get("EDGAR Country Code") or "").strip().upper()
            if not code:
                continue
            years = per_country.setdefault(code, {})
            for col_name, raw_value in row.items():
                if col_name in ("EDGAR Country Code", "Country", "Region"):
                    continue
                m = year_col_re.match(col_name)
                if not m:
                    continue
                field, year_str = m.group(1), int(m.group(2))
                bucket = years.setdefault(year_str, {})
                if field.startswith("Rank"):
                    bucket[field] = (raw_value or "").strip()
                else:
                    bucket[field] = parse_float(raw_value)

    def build_entry(years_dict):
        """Given {year: {field: value}} for one country, build snapshot + trends."""
        if not years_dict:
            return None
        all_years = sorted(years_dict)
        years_with_data = [
            y for y in all_years
            if years_dict[y].get("Global_Emissions") is not None
               or years_dict[y].get("Transport_Emissions") is not None
        ]
        latest_year = years_with_data[-1] if years_with_data else all_years[-1]

        latest = years_dict[latest_year]
        total = latest.get("Global_Emissions")
        transport = latest.get("Transport_Emissions")
        share = latest.get("Share_Transport_National")
        if share is not None:
            share = round(share * 100, 2)  # ghg.csv stores fraction, profiles use %
        global_share = latest.get("Share_Transport_Global")
        if global_share is not None:
            global_share = round(global_share * 100, 4)
        per_capita = latest.get("PerCapita_Transport")

        # Sector rank: only set if "Transport" appears in Rank1/2/3_Sector
        rank = None
        for i in (1, 2, 3):
            if latest.get(f"Rank{i}_Sector") == "Transport":
                rank = i
                break

        snapshot = {
            "total_mt": round(total, 2) if total is not None else None,
            "transport_mt": round(transport, 2) if transport is not None else None,
            "transport_share_pct": share,
            "transport_global_share_pct": global_share,
            "transport_per_capita": round(per_capita, 3) if per_capita is not None else None,
            "transport_sector_rank": rank,
            "year": latest_year,
            "source": "EDGAR",
        }

        trend_years, trend_total, trend_transport = [], [], []
        for y in all_years:
            t = years_dict[y].get("Global_Emissions")
            tr = years_dict[y].get("Transport_Emissions")
            if t is None and tr is None:
                continue
            trend_years.append(y)
            trend_total.append(round(t, 2) if t is not None else None)
            trend_transport.append(round(tr, 2) if tr is not None else None)

        trends = {
            "years": trend_years,
            "total": trend_total,
            "transport": trend_transport,
            "source": "EDGAR",
        } if trend_years else None

        return {"snapshot": snapshot, "trends": trends}

    result = {}
    for code, years_dict in per_country.items():
        entry = build_entry(years_dict)
        if entry:
            result[code] = entry

    # ── EU (EEU) aggregate, summed from member states ──────────────────
    if EU_CODE not in result:
        eu_years = defaultdict(lambda: [0.0, 0.0, 0])  # year -> [total, transport, n_reporting]
        for c in EU_MEMBER_ISO3:
            member_years = per_country.get(c, {})
            for y, fields in member_years.items():
                t = fields.get("Global_Emissions")
                if t is None:
                    continue
                tr = fields.get("Transport_Emissions") or 0
                eu_years[y][0] += t
                eu_years[y][1] += tr
                eu_years[y][2] += 1

        eu_trend_years, eu_total, eu_transport = [], [], []
        for y in sorted(eu_years):
            tot, tra, n = eu_years[y]
            if n < 20:  # only years where most member states reported
                continue
            eu_trend_years.append(y)
            eu_total.append(round(tot, 2))
            eu_transport.append(round(tra, 2))

        if eu_trend_years:
            latest_y = eu_trend_years[-1]
            latest_tot = eu_total[-1]
            latest_tra = eu_transport[-1]
            eu_share = round(latest_tra / latest_tot * 100, 2) if latest_tot else None
            result[EU_CODE] = {
                "snapshot": {
                    "total_mt": latest_tot,
                    "transport_mt": latest_tra,
                    "transport_share_pct": eu_share,
                    "transport_global_share_pct": None,
                    "transport_per_capita": None,
                    "transport_sector_rank": None,
                    "year": latest_y,
                    "source": "EDGAR (sum of 27 member states)",
                },
                "trends": {
                    "years": eu_trend_years,
                    "total": eu_total,
                    "transport": eu_transport,
                    "source": "EDGAR (sum of 27 member states)",
                },
            }

    populated = sum(
        1 for v in result.values()
        if any(v["snapshot"].get(k) is not None
               for k in ("total_mt", "transport_mt", "transport_share_pct"))
    )
    print(f"   ✅  GHG data loaded from {path.name}: {len(result)} countries ({populated} with data)")
    return result


# ============================================================================
# Processing
# ============================================================================

def process(excel_path, ghg_by_country=None):
    print(f"📊  Opening {excel_path} …")
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    ghg_by_country = ghg_by_country or {}

    # ------------------------------------------------------------------ Country
    print("🌍  Reading Country sheet …")
    countries = {}
    for r in rows_as_dicts(wb["Country"]):
        code = clean(r.get("Country Codes"))
        name = clean(r.get("Country"))
        if not code or not name:
            continue
        coalitions = []
        for col, label in COALITION_COLUMNS.items():
            # tolerate trailing-space variants in headers
            val = r.get(col) or r.get(col + " ")
            if clean(val):
                coalitions.append(label)
        # Emissions: use ghg.csv (via load_ghg_csv) as primary source;
        # fall back to Excel columns if a country is missing from EDGAR.
        _ghg = ghg_by_country.get(code, {}) if ghg_by_country else {}
        total     = _ghg.get("total_mt")     or r.get("GHG total 2023 (Mt)")
        transport = _ghg.get("transport_mt") or r.get("GHG transport 2023 (Mt)")
        share = _ghg.get("transport_share_pct")
        if share is None and isinstance(total, (int, float)) and isinstance(transport, (int, float)) and total > 0:
            share = round(transport / total * 100, 2)
        _ghg_year   = _ghg.get("year")   or 2023
        _ghg_source = _ghg.get("source") or "EDGAR"
        countries[code] = {
            "code": code,
            "iso2": iso2_for(code),
            "name": name,
            "slug": slugify(name),
            "annex": clean(r.get("Annex I or Non-Annex I")),
            "region": clean(r.get("Region")),
            "income": clean(r.get("Income Level Group")),
            "memberships": {
                "g20": yesno(r.get("G20")) or False,
                "g7": yesno(r.get("G7")) or False,
                "oecd": clean(r.get("OECD")) == "OECD",
                "eu27": clean(r.get("EU27")) == "EU" or code in EU_MEMBER_ISO3,
                "mena": clean(r.get("MENA region")) == "MENA"
                        or clean(r.get("MENA region ")) == "MENA",
            },
            "coalitions": coalitions,
            "net_zero_target": yesno(r.get("Net-zero target")) or False,
            "ice_phaseout": {
                "has": yesno(r.get("ICE phase-out target")) or False,
                "type": clean(r.get("ICE phase-out target type")),
                "year": clean(r.get("ICE phase-out target year")),
            },
            "emissions": {
                "year": _ghg_year,
                "source": _ghg_source,
                "total_mt": round(total, 2) if isinstance(total, (int, float)) else None,
                "transport_mt": round(transport, 2) if isinstance(transport, (int, float)) else None,
                "transport_share_pct": share,
                "transport_global_share_pct": _ghg.get("transport_global_share_pct"),
                "transport_per_capita": _ghg.get("transport_per_capita"),
                "transport_sector_rank": _ghg.get("transport_sector_rank"),
            },
            "eu_member": code in EU_MEMBER_ISO3,
        }

    # The EU row may exist without emissions data — backfill from member sums.
    if EU_CODE in countries and countries[EU_CODE]["emissions"]["total_mt"] is None:
        eu_total = sum(countries[c]["emissions"]["total_mt"] or 0
                       for c in EU_MEMBER_ISO3 if c in countries)
        eu_transport = sum(countries[c]["emissions"]["transport_mt"] or 0
                           for c in EU_MEMBER_ISO3 if c in countries)
        if eu_total:
            countries[EU_CODE]["emissions"].update({
                "source": "EDGAR (sum of 27 member states)",
                "total_mt": round(eu_total, 2),
                "transport_mt": round(eu_transport, 2),
                "transport_share_pct": round(eu_transport / eu_total * 100, 2),
            })

    # The EU itself is not in the Country sheet — create a collective entry.
    if EU_CODE not in countries:
        eu_total = sum(
            countries[c]["emissions"]["total_mt"] or 0
            for c in EU_MEMBER_ISO3 if c in countries
        )
        eu_transport = sum(
            countries[c]["emissions"]["transport_mt"] or 0
            for c in EU_MEMBER_ISO3 if c in countries
        )
        countries[EU_CODE] = {
            "code": EU_CODE, "iso2": "eu", "name": "European Union",
            "slug": "european-union",
            "annex": "Annex I", "region": "Europe", "income": "High-income",
            "memberships": {"g20": True, "g7": False, "oecd": False,
                            "eu27": True, "mena": False},
            "coalitions": [],
            "net_zero_target": True,
            "ice_phaseout": {"has": False, "type": None, "year": None},
            "emissions": {
                "year": 2023, "source": "EDGAR (sum of 27 member states)",
                "total_mt": round(eu_total, 2) if eu_total else None,
                "transport_mt": round(eu_transport, 2) if eu_transport else None,
                "transport_share_pct": round(eu_transport / eu_total * 100, 2)
                                       if eu_total else None,
            },
            "eu_member": False,
        }

    # ------------------------------------------------------------------ Document
    print("📄  Reading Document sheet …")
    docs_by_country = defaultdict(list)
    doc_meta = {}  # Document ID → summary (for joining other sheets)
    for r in rows_as_dicts(wb["Document"]):
        code = clean(r.get("Country Code"))
        if not code:
            continue
        doc = {
            "id": r.get("Document ID"),
            "type": clean(r.get("Type of document")),
            "name": clean(r.get("Document name")),
            "version": clean(r.get("Version number")),
            "generation": _profiles_get_gen(r.get("Version number")),
            "date": excel_date(r.get("Date")),
            "status": clean(r.get("Status")),
            "url": clean(r.get("URL")),
            "transport": {
                "has_content": yesno(r.get("Transport content")),
                "mitigation_target": yesno(r.get("Contains transport mitigation target")),
                "target_summary": clean(r.get("Summary transport target")),
                "ghg_target_type": clean(r.get("GHG transport target type")),
                "adaptation_target": yesno(r.get("Contains transport adaptation target")),
                "mitigation_measures": yesno(r.get("Contains transport mitigation measures")),
                "adaptation_measures": yesno(r.get("Contains transport adaptation measures")),
                "benefits": yesno(r.get("Contains benefits")),
                "just_transition": yesno(r.get("Contains reference to just transition")),
            },
            "counts": {"targets": 0, "measures": 0, "adaptation": 0, "benefits": 0},
        }
        docs_by_country[code].append(doc)
        doc_meta[doc["id"]] = (code, doc)

    # ------------------------------------------------------------------ Targets
    print("🎯  Reading Targets sheet …")
    targets_by_country = defaultdict(list)
    for r in rows_as_dicts(wb["Targets"]):
        code = clean(r.get("Country Code"))
        if not code:
            continue
        t = {
            "doc_id": r.get("Document ID"),
            "document": clean(r.get("Document name")),
            "version": clean(r.get("Version number")),
            "doc_type": clean(r.get("Type of Document")),
            "status": clean(r.get("Status")),
            "area": clean(r.get("Target area")),
            "scope": clean(r.get("Target scope")),
            "ghg": clean(r.get("GHG target?")),
            "type": clean(r.get("Target type")),
            "conditionality": clean(r.get("Conditionality")),
            "year": clean(r.get("Target Year")),
            "content": clean(r.get("Content")),
            "page": clean(r.get("Page Number")),
        }
        targets_by_country[code].append(t)
        if t["doc_id"] in doc_meta:
            doc_meta[t["doc_id"]][1]["counts"]["targets"] += 1

    # ------------------------------------------------------------------ Mitigation
    print("🚲  Reading Mitigation sheet …")
    measures_by_country = defaultdict(list)
    for r in rows_as_dicts(wb["Mitigation"]):
        code = clean(r.get("Country Code"))
        if not code:
            continue
        modes = [m for m in MODE_COLUMNS if clean(r.get(m))]
        geos = [g for g in GEO_COLUMNS if clean(r.get(g))]
        m = {
            "doc_id": r.get("Document ID"),
            "document": clean(r.get("Document name")),
            "version": clean(r.get("Version number")),
            "doc_type": clean(r.get("Type of Document")),
            "status": clean(r.get("Status")),
            "category": clean(r.get("Category")),
            "purpose": clean(r.get("Purpose")),
            "instrument": clean(r.get("Instrument")),
            "asi": norm_asi(r.get("A-S-I")),
            "quote": clean(r.get("Quote")),
            "page": clean(r.get("Page Number")),
            "modes": modes,
            "geography": geos,
            "measure_status": clean(r.get("Status of measure")),
        }
        measures_by_country[code].append(m)
        if m["doc_id"] in doc_meta:
            doc_meta[m["doc_id"]][1]["counts"]["measures"] += 1

    # ------------------------------------------------------------------ Adaptation
    print("🌊  Reading Adaptation sheet …")
    adaptation_by_country = defaultdict(list)
    for r in rows_as_dicts(wb["Adaptation"]):
        code = clean(r.get("Country Code"))
        if not code:
            continue
        modes = [m for m in MODE_COLUMNS if clean(r.get(m))]
        a = {
            "doc_id": r.get("Document ID"),
            "document": clean(r.get("Document name")),
            "version": clean(r.get("Version number")),
            "doc_type": clean(r.get("Type of Document")),
            "status": clean(r.get("Status")),
            "category": clean(r.get("Category")),
            "measure": clean(r.get("Measure")),
            "quote": clean(r.get("Quote")),
            "page": clean(r.get("Page Number")),
            "modes": modes,
        }
        adaptation_by_country[code].append(a)
        if a["doc_id"] in doc_meta:
            doc_meta[a["doc_id"]][1]["counts"]["adaptation"] += 1

    # ------------------------------------------------------------------ Benefits
    print("💚  Reading Benefits sheet …")
    benefits_by_country = defaultdict(list)
    for r in rows_as_dicts(wb["Benefits"]):
        code = clean(r.get("Country Code"))
        if not code:
            continue
        b = {
            "doc_id": r.get("Document ID"),
            "document": clean(r.get("Document name")),
            "version": clean(r.get("Version number")),
            "status": clean(r.get("Status")),
            "type": clean(r.get("Type of benefit")),
            "quote": clean(r.get("Quote")),
            "page": clean(r.get("Page Number")),
        }
        benefits_by_country[code].append(b)
        if b["doc_id"] in doc_meta:
            doc_meta[b["doc_id"]][1]["counts"]["benefits"] += 1

    # ------------------------------------------------------------------ References
    print("🔗  Reading References sheet …")
    references_by_country = defaultdict(list)
    for r in rows_as_dicts(wb["References"]):
        code = clean(r.get("Country Code"))
        if not code:
            continue
        references_by_country[code].append({
            "doc_id": r.get("Document ID"),
            "document": clean(r.get("Document name")),
            "status": clean(r.get("Status")),
            "further_type": clean(r.get("Further document type")),
            "url": clean(r.get("URL to further document")),
            "quote": clean(r.get("Quote")),
            "page": clean(r.get("Page Number")),
        })

    wb.close()
    return (countries, docs_by_country, targets_by_country, measures_by_country,
            adaptation_by_country, benefits_by_country, references_by_country)


# ============================================================================
# Derived data: target years, similarity
# ============================================================================

def extract_target_years(targets):
    """Years of active GHG targets, by scope, for the 'where targets point' axis."""
    seen = set()
    out = []
    for t in targets:
        if t["status"] != "Active" or t["ghg"] != "GHG":
            continue
        try:
            y = int(str(t["year"])[:4])
        except (TypeError, ValueError):
            continue
        if not 2000 <= y <= 2100:
            continue
        area = t["area"] or ""
        if "Transport sector mitigation" in area:
            scope = "transport"
        elif "Net zero" in area:
            scope = "net-zero"
        elif "Overall mitigation" in area:
            scope = "economy-wide"
        else:
            continue
        if (y, scope) in seen:
            continue
        seen.add((y, scope))
        out.append({"year": y, "scope": scope})
    return sorted(out, key=lambda x: x["year"])


def cosine(a, b, keys):
    import math
    dot = sum(a.get(k, 0) * b.get(k, 0) for k in keys)
    na = math.sqrt(sum(a.get(k, 0) ** 2 for k in keys))
    nb = math.sqrt(sum(b.get(k, 0) ** 2 for k in keys))
    return dot / (na * nb) if na and nb else 0.0


def compute_similar(profiles, max_each=6):
    """Peer discovery — deliberately NOT a ranking. Three neutral lenses:
    same region (alphabetical), similar transport share of emissions
    (closest |delta|), and 'betting on the same priorities' (cosine
    similarity of the active-measure category mix, min. 3 measures)."""
    cats = set()
    for p in profiles.values():
        cats.update(p["category_summary"].keys())
    cats = sorted(cats)

    def chip(q, extra=None):
        c = {"code": q["code"], "name": q["name"], "iso2": q["iso2"]}
        if extra:
            c.update(extra)
        return c

    for code, p in profiles.items():
        others = [q for q in profiles.values()
                  if q["code"] != code and q["documents"]]

        region = sorted(
            (q for q in others if q["region"] and q["region"] == p["region"]),
            key=lambda q: q["name"])[:max_each]

        share = p["emissions"]["transport_share_pct"]
        emissions = []
        if share is not None:
            cands = [q for q in others
                     if q["emissions"]["transport_share_pct"] is not None]
            cands.sort(key=lambda q:
                       abs(q["emissions"]["transport_share_pct"] - share))
            emissions = [chip(q, {"share": q["emissions"]["transport_share_pct"]})
                         for q in cands[:max_each]]

        priorities = []
        if sum(p["category_summary"].values()) >= 3:
            scored = []
            for q in others:
                # EU members share the EU NDC's measures, so member-to-member
                # 'priorities' similarity is an artifact of the same document
                # — not an insight. Compare EU members with non-EU peers only.
                eu_side_p = p["eu_member"] or p["code"] == EU_CODE
                eu_side_q = q["eu_member"] or q["code"] == EU_CODE
                if eu_side_p and eu_side_q:
                    continue
                if sum(q["category_summary"].values()) < 3:
                    continue
                s = cosine(p["category_summary"], q["category_summary"], cats)
                if s >= 0.6:
                    shared = max(
                        (k for k in cats if p["category_summary"].get(k)
                         and q["category_summary"].get(k)),
                        key=lambda k: min(p["category_summary"][k],
                                          q["category_summary"][k]),
                        default=None)
                    scored.append((s, q, shared))
            scored.sort(key=lambda x: -x[0])
            priorities = [chip(q, {"shared_focus": shared})
                          for s, q, shared in scored[:max_each]]

        p["similar"] = {
            "region": [chip(q) for q in region],
            "emissions": emissions,
            "priorities": priorities,
        }


# ============================================================================
# Profile assembly
# ============================================================================

def sort_docs(docs):
    type_order = {"NDC": 0, "LTS": 1, "BTR": 2,
                  "National policy document": 3, "Other": 4}
    return sorted(docs, key=lambda d: (
        type_order.get(d["type"], 9), d["date"] or "0000-00-00", str(d["version"] or "")))


def summarise_active(measures):
    """A-S-I and category summaries for measures in active documents."""
    asi = defaultdict(int)
    cats = defaultdict(int)
    for m in measures:
        if m["status"] != "Active":
            continue
        for a in m["asi"]:
            asi[a] += 1
        if m["category"]:
            cats[m["category"]] += 1
    return (
        {k: asi[k] for k in ("Avoid", "Shift", "Improve") if asi.get(k)},
        dict(sorted(cats.items(), key=lambda kv: -kv[1])),
    )


def build_profiles(data, publications, ghg_by_country=None):
    (countries, docs_by_country, targets_by_country, measures_by_country,
     adaptation_by_country, benefits_by_country, references_by_country) = data

    profiles = {}
    eu_docs = sort_docs(docs_by_country.get(EU_CODE, []))
    eu_ndc_docs = [d for d in eu_docs if d["type"] == "NDC"]

    for code, base in countries.items():
        docs = sort_docs(docs_by_country.get(code, []))
        is_eu_member = base["eu_member"]
        reports_via_eu = is_eu_member and any(
            d["status"] in EU_STATUSES for d in docs
        ) or (is_eu_member and bool(eu_ndc_docs))

        # EU members: their own NDC rows with "Covered by EU" status are
        # placeholders — replace them with the collective EU NDC documents.
        merged_docs = [d for d in docs if d["status"] not in EU_STATUSES] \
            if is_eu_member else list(docs)
        if is_eu_member:
            for d in eu_ndc_docs:
                dd = dict(d)
                dd["via_eu"] = True
                merged_docs.append(dd)
            merged_docs = sort_docs(merged_docs)

        targets = targets_by_country.get(code, [])
        measures = measures_by_country.get(code, [])
        adaptation = adaptation_by_country.get(code, [])
        benefits = benefits_by_country.get(code, [])
        references = references_by_country.get(code, [])

        # EU members inherit the EU-level entries for their collective NDC
        if is_eu_member:
            for src, dst in (
                (targets_by_country.get(EU_CODE, []), targets),
                (measures_by_country.get(EU_CODE, []), measures),
                (adaptation_by_country.get(EU_CODE, []), adaptation),
                (benefits_by_country.get(EU_CODE, []), benefits),
            ):
                for item in src:
                    it = dict(item)
                    it["via_eu"] = True
                    dst.append(it)

        asi_summary, cat_summary = summarise_active(measures)

        active_ndc = next((d for d in merged_docs
                           if d["type"] == "NDC" and d["status"] == "Active"), None)
        active_lts = next((d for d in merged_docs
                           if d["type"] == "LTS" and d["status"] == "Active"), None)

        profile = dict(base)
        profile.update({
            "reports_via_eu": bool(reports_via_eu),
            "documents": merged_docs,
            "active": {
                "ndc_version": active_ndc["version"] if active_ndc else None,
                "lts": bool(active_lts),
            },
            "targets": targets,
            "measures": measures,
            "adaptation": adaptation,
            "benefits": benefits,
            "references": references,
            "asi_summary": asi_summary,
            "category_summary": cat_summary,
            "target_years": extract_target_years(targets),
            # BTR integration lands with the next database release; the schema
            # slot exists so the front end and downstream users can rely on it.
            "btr": {"available": False, "documents": [], "measures": []},
            "links": {
                "changing_transport_search":
                    f"https://changing-transport.org/?s={base['name'].replace(' ', '+')}",
                "tdc_search":
                    "https://portal.transport-data.org/search"
                    + (f"?country={base['iso2'].upper()}" if base["iso2"] else ""),
            },
            "publications": publications.get(code, []),
        })
        profiles[code] = profile

    return profiles


STATIC_DIR = Path("profiles/countries")


def build_static_pages(profiles, template="profiles/country.html"):
    """Generate a pre-rendered shell per country at countries/<slug>/index.html.
    Each shell carries its own <title> and meta description (sharing / SEO)
    and bootstraps the same renderer with the country code baked in.
    Re-generated on every database push, so 'static' stays current."""
    tpl_path = Path(template)
    if not tpl_path.exists():
        return 0
    tpl = tpl_path.read_text(encoding="utf-8")
    n = 0
    for code, p in profiles.items():
        title = f"{p['name']} — Transport in Climate Policy | GIZ-SLOCAT Transport Tracker"
        e = p["emissions"]
        bits = [f"How {p['name']}'s climate documents address transport: "
                f"{len([d for d in p['documents'] if d['type'] == 'NDC'])} NDC submissions"]
        if p["active"]["lts"]:
            bits.append("a long-term strategy")
        if e["transport_share_pct"] is not None:
            bits.append(f"transport is {e['transport_share_pct']}% of emissions")
        desc = ", ".join(bits) + ". GIZ-SLOCAT NDC Transport Tracker."
        desc = desc.replace('"', "'")

        html = tpl.replace(
            "<title>Country Profile — GIZ-SLOCAT Transport Tracker</title>",
            f"<title>{title}</title>\n"
            f'<meta name="description" content="{desc}">')
        html = (html
                .replace('href="../assets/design-tokens.css"',
                         'href="../../../assets/design-tokens.css"')
                .replace('href="styles.css"', 'href="../../styles.css"')
                .replace('src="js/country.js"', 'src="../../js/country.js"')
                .replace('href="index.html"', 'href="../../index.html"'))
        html = html.replace(
            "</head>",
            f'<script>window.CP_CODE="{code}";window.CP_BASE="../../";</script>\n</head>')

        out = STATIC_DIR / p["slug"]
        out.mkdir(parents=True, exist_ok=True)
        (out / "index.html").write_text(html, encoding="utf-8")
        n += 1
    return n


def build_index(profiles):
    items = []
    for code, p in sorted(profiles.items(), key=lambda kv: kv[1]["name"]):
        has_transport = any(
            d["status"] == "Active" and d["transport"]["has_content"]
            for d in p["documents"]
        )
        items.append({
            "code": code,
            "iso2": p["iso2"],
            "name": p["name"],
            "region": p["region"],
            "income": p["income"],
            "eu_member": p["eu_member"],
            "ndc_version": p["active"]["ndc_version"],
            "has_lts": p["active"]["lts"],
            "has_transport_content": has_transport,
            "n_documents": len(p["documents"]),
            "n_measures": len(p["measures"]),
            "transport_share_pct": p["emissions"]["transport_share_pct"],
        })
    return items


# ============================================================================
# Main
# ============================================================================



def run_profiles(excel_path):
    """Country profile JSONs + static pages (profiles/)."""
    # ── Publications ──────────────────────────────────────────────────
    pubs_path = Path("data/publications.json")
    publications = {}
    if pubs_path.exists():
        pub_data = json.loads(pubs_path.read_text(encoding="utf-8"))
        publications = pub_data.get("by_country", {})
        global_pubs = publications.get("GLOBAL", [])
        n_countries = len([k for k in publications if k != "GLOBAL"])
        print(f"   ✅  Publications loaded: {n_countries} countries + {len(global_pubs)} GLOBAL entries")
    else:
        global_pubs = []
        print("   ⚠  data/publications.json not found — run scripts/build_data_files.py")

    # Merge GLOBAL publications into every country
    for code in list(publications.keys()):
        if code != "GLOBAL":
            publications[code] = publications[code] + global_pubs

    # ── GHG data (single source of truth: data/ghg.csv) ────────────────
    ghg_data = load_ghg_csv()
    ghg_by_country = {code: v["snapshot"] for code, v in ghg_data.items()}

    data = process(excel_path, ghg_by_country=ghg_by_country)
    profiles = build_profiles(data, publications, ghg_by_country=ghg_by_country)
    compute_similar(profiles)

    trends = {code: v["trends"] for code, v in ghg_data.items() if v["trends"]}
    today = str(date.today())
    for code, profile in profiles.items():
        profile["trends"] = trends.get(code)
        profile["meta"] = {"generated": today, "database": Path(excel_path).name}

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    for code, profile in profiles.items():
        (OUT_DIR / f"{code}.json").write_text(
            json.dumps(profile, ensure_ascii=False, separators=(",", ":")),
            encoding="utf-8")

    index = {
        "metadata": {"generated": today,
                     "source": "GIZ-SLOCAT Transport Tracker database",
                     "countries": len(profiles)},
        "countries": build_index(profiles),
    }
    (OUT_DIR / "index.json").write_text(
        json.dumps(index, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8")

    n_static = build_static_pages(profiles)
    print(f"💾  {len(profiles)} country profiles + index.json → {OUT_DIR}/")
    print(f"💾  {n_static} static pages → profiles/countries/<slug>/")
    return profiles


# ============================================================================
# Entry point — one pipeline, three products
# ============================================================================

def main():
    print("=" * 70)
    print("GIZ-SLOCAT Transport Tracker — Ecosystem Data Update")
    print("=" * 70)

    if len(sys.argv) > 1:
        excel_path = Path(sys.argv[1])
    else:
        # Exact expected filename — deliberately NOT a glob. data/ also holds
        # publications.xlsx (and possibly other .xlsx in the future); picking
        # by glob+sort could silently process the wrong workbook.
        excel_path = Path("data/GIZ-SLOCAT_Transport-Tracker-database.xlsx")
        if not excel_path.exists():
            print("❌  data/GIZ-SLOCAT_Transport-Tracker-database.xlsx not found.")
            print("    Filename must be exact (case-sensitive). Either commit it")
            print("    there (Option A) or run pipeline/fetch_database.py (Option B).")
            return 1

    if not excel_path.exists():
        print(f"❌  Excel file not found: {excel_path}")
        return 1

    print(f"📂  Database: {excel_path}\n")

    # Validate schema ONCE up front — covers all three parse passes below
    _wb = openpyxl.load_workbook(excel_path, read_only=True)
    validate_workbook(_wb)
    _wb.close()

    run_dashboards(excel_path)
    print()
    run_profiles(excel_path)

    print("\n" + "=" * 70)
    print("✅  Ecosystem updated: main dashboard · comparison · country profiles")
    print("=" * 70)
    return 0


if __name__ == "__main__":
    sys.exit(main())