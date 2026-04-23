#!/usr/bin/env python3
"""
NDC Transport Tracker - Data Processing Script
Processes the GIZ-SLOCAT Excel database and generates:
  - data/processed/data.json
"""

import json
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path

import openpyxl


# ============================================================================
# Helpers
# ============================================================================

def get_gen(version):
    """Map a version string like 'NDC 2.1' to 'gen1' / 'gen2' / 'gen3'."""
    v = str(version).strip()
    if v.startswith("NDC 1"): return "gen1"
    if v.startswith("NDC 2"): return "gen2"
    if v.startswith("NDC 3"): return "gen3"
    return None


# ============================================================================
# Main processing
# ============================================================================

def process_excel(excel_path):
    print("📊  Opening Excel file…")
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)

    doc_sheet     = wb["Document"]
    mit_sheet     = wb["Mitigation"]
    targets_sheet = wb["Targets"]

    EU_STATUSES = {"Covered by EU", "Covered by EU archived"}

    # Column positions (0-indexed)
    D_DOCID     = 0
    D_CODE      = 1
    D_NAME      = 2
    D_TYPE      = 4
    D_VERSION   = 7
    D_STATUS    = 9
    D_TRANSPORT = 10
    D_REGION    = 25

    M_DOCID    = 0
    M_CODE     = 1
    M_VERSION  = 6
    M_CATEGORY = 9

    # Targets sheet columns
    T_DOCID       = 0
    T_CODE        = 1
    T_TYPE        = 3
    T_TARGET_AREA = 8
    T_GHG_FLAG    = 10

    # ── Pass 1: Document sheet ─────────────────────────────────────────
    country_master = {}
    doc_id_info    = {}

    for row in doc_sheet.iter_rows(min_row=2, values_only=True):
        if not row[D_DOCID]: break

        doc_id  = row[D_DOCID]
        code    = row[D_CODE]
        name    = row[D_NAME]
        dtype   = row[D_TYPE]
        version = row[D_VERSION]
        status  = row[D_STATUS]
        has_t   = row[D_TRANSPORT]
        region  = row[D_REGION]

        if dtype != "NDC" or not code or not version: continue
        if status in EU_STATUSES: continue

        gen = get_gen(version)
        if not gen: continue

        code   = str(code).strip()
        status = str(status).strip()

        if code not in country_master:
            country_master[code] = {
                "name":   str(name).strip() if name else code,
                "region": str(region).strip() if region else "Unknown",
            }

        doc_id_info[doc_id] = {
            "code":      code,
            "gen":       gen,
            "status":    status,
            "has_t":     str(has_t).strip().lower() == "yes",
            "region":    str(region).strip() if region else "Unknown",
        }

    print(f"   ✓ {len(country_master)} countries, {len(doc_id_info)} NDC documents")

    # ── Latest active doc per country ──────────────────────────────────
    GEN_PRIORITY = {"gen1": 1, "gen2": 2, "gen3": 3}
    country_best = {}

    for doc_id, info in doc_id_info.items():
        if info["status"] != "Active": continue
        code = info["code"]
        prio = GEN_PRIORITY[info["gen"]]
        if code not in country_best or prio > country_best[code][0]:
            country_best[code] = (prio, doc_id)

    latest_doc_ids = {v[1] for v in country_best.values()}
    latest_gen_map = {
        code: ["gen1", "gen2", "gen3"][prio - 1]
        for code, (prio, _) in country_best.items()
    }

    # ── Per-gen active/archived country counts ─────────────────────────
    gen_active_codes   = {g: set() for g in ["gen1", "gen2", "gen3"]}
    gen_archived_codes = {g: set() for g in ["gen1", "gen2", "gen3"]}

    for info in doc_id_info.values():
        code = info["code"]; gen = info["gen"]
        if info["status"] == "Active":
            gen_active_codes[gen].add(code)
        else:
            gen_archived_codes[gen].add(code)

    gen_counts = {
        g: {
            "active":   len(gen_active_codes[g]),
            "archived": len(gen_archived_codes[g]),
        }
        for g in ["gen1", "gen2", "gen3"]
    }
    gen_counts["latest"] = {"active": len(latest_doc_ids), "archived": 0}

    # ── Targets sheet: identify transport and GHG transport doc_ids ────
    print("🔄  Scanning Targets sheet…")

    doc_has_transport     = set()  # any transport sector mitigation target
    doc_has_ghg_transport = set()  # quantified GHG transport target

    for row in targets_sheet.iter_rows(min_row=2, values_only=True):
        if not row[T_DOCID]: break

        doc_id      = row[T_DOCID]
        doc_type    = row[T_TYPE]
        target_area = str(row[T_TARGET_AREA] or "")
        ghg_flag    = str(row[T_GHG_FLAG] or "")

        if doc_type != "NDC": continue
        if doc_id not in doc_id_info: continue

        if "Transport sector mitigation" in target_area:
            doc_has_transport.add(doc_id)
            if ghg_flag == "GHG":
                doc_has_ghg_transport.add(doc_id)

    print(f"   ✓ {len(doc_has_transport)} docs with transport target, "
          f"{len(doc_has_ghg_transport)} with GHG transport target")

    # ── TAB 1: per-generation stats + per-country data ─────────────────
    print("🔄  Computing Tab 1 data…")

    # Per gen: collect unique countries and their transport flags
    country_gen_docs = defaultdict(lambda: defaultdict(list))
    for doc_id, info in doc_id_info.items():
        country_gen_docs[info["code"]][info["gen"]].append(doc_id)

    GEN_META = {
        "gen1": {"name": "First Generation",  "period": "2015–2019"},
        "gen2": {"name": "Second Generation", "period": "2020–2024"},
        "gen3": {"name": "Third Generation",  "period": "2024–ongoing"},
    }

    tab1_generations = {}
    for gen in ["gen1", "gen2", "gen3"]:
        submitted        = set()
        with_transport   = set()
        with_ghg         = set()
        reg_submitted    = defaultdict(set)
        reg_transport    = defaultdict(set)
        reg_ghg          = defaultdict(set)

        for code, gen_docs in country_gen_docs.items():
            if gen not in gen_docs: continue
            docs   = gen_docs[gen]
            region = doc_id_info[docs[0]]["region"]
            submitted.add(code)
            reg_submitted[region].add(code)

            for doc_id in docs:
                if doc_id in doc_has_transport:
                    with_transport.add(code)
                    reg_transport[region].add(code)
                if doc_id in doc_has_ghg_transport:
                    with_ghg.add(code)
                    reg_ghg[region].add(code)

        regions_out = {
            reg: {
                "total":              len(reg_submitted[reg]),
                "with_transport":     len(reg_transport.get(reg, set())),
                "with_ghg_transport": len(reg_ghg.get(reg, set())),
            }
            for reg in reg_submitted
        }

        tab1_generations[gen] = {
            **GEN_META[gen],
            "total_submitted":     len(submitted),
            "with_transport":      len(with_transport),
            "with_ghg_transport":  len(with_ghg),
            "regions":             regions_out,
        }

    # Per-country: latest_has_transport, latest_has_ghg_transport
    countries_tab1 = {}

    for row in doc_sheet.iter_rows(min_row=2, values_only=True):
        if not row[D_DOCID]: break

        code    = row[D_CODE]
        dtype   = row[D_TYPE]
        version = row[D_VERSION]
        status  = row[D_STATUS]
        has_t   = row[D_TRANSPORT]
        region  = row[D_REGION]

        if dtype != "NDC" or not code or not version: continue
        if status in EU_STATUSES: continue

        gen = get_gen(version)
        if not gen: continue

        code    = str(code).strip()
        has_t_b = str(has_t or "").strip().lower() == "yes"
        region  = str(region or "Unknown").strip()

        if code not in countries_tab1:
            countries_tab1[code] = {
                "iso3":                    code,
                "name":                    country_master[code]["name"],
                "region":                  country_master[code]["region"],
                "latest_active_gen":       latest_gen_map.get(code),
                "latest_has_transport":    None,
                "latest_has_ghg_transport": False,
                "generations":             {},
            }

        gd = countries_tab1[code]["generations"]
        if gen not in gd:
            gd[gen] = {"has_transport": has_t_b, "version": str(version).strip()}
        elif has_t_b:
            gd[gen]["has_transport"] = True

    # Fill latest transport flags from latest active doc
    for code, (prio, doc_id) in country_best.items():
        if code not in countries_tab1: continue
        lat_gen = latest_gen_map.get(code)
        cd = countries_tab1[code]
        if lat_gen and lat_gen in cd["generations"]:
            cd["latest_has_transport"] = cd["generations"][lat_gen]["has_transport"]
        cd["latest_has_ghg_transport"] = doc_id in doc_has_ghg_transport

    print(f"   ✓ Tab1 built for {len(countries_tab1)} countries")

    # ── TAB 2: mitigation measures ─────────────────────────────────────
    print("🔄  Computing Tab 2 data…")

    cat_lgb_sets = defaultdict(lambda: {
        "gen1": set(), "gen2": set(), "gen3": set(), "mentions": 0
    })

    cat_gen_aa_sets = {
        g: defaultdict(lambda: {
            "active_c": set(), "archived_c": set(),
            "active_m": 0,     "archived_m": 0,
        })
        for g in ["gen1", "gen2", "gen3"]
    }

    country_gen_cats    = {}
    country_latest_cats = {}

    for row in mit_sheet.iter_rows(min_row=2, values_only=True):
        if not row[M_DOCID]: break

        doc_id   = row[M_DOCID]
        code     = row[M_CODE]
        version  = row[M_VERSION]
        category = row[M_CATEGORY]

        if not category or not code or not version or not doc_id: continue
        if doc_id not in doc_id_info: continue

        code = str(code).strip()
        cat  = str(category).strip()
        gen  = get_gen(version)
        if not gen: continue

        info   = doc_id_info[doc_id]
        status = info["status"]

        if doc_id in latest_doc_ids:
            lat_gen = latest_gen_map.get(code)
            if lat_gen:
                cat_lgb_sets[cat][lat_gen].add(code)
                cat_lgb_sets[cat]["mentions"] += 1
            if code not in country_latest_cats:
                country_latest_cats[code] = {}
            country_latest_cats[code][cat] = country_latest_cats[code].get(cat, 0) + 1

        s = cat_gen_aa_sets[gen][cat]
        if status == "Active":
            s["active_c"].add(code); s["active_m"] += 1
        else:
            s["archived_c"].add(code); s["archived_m"] += 1

        if code not in country_gen_cats:
            country_gen_cats[code] = {}
        if gen not in country_gen_cats[code]:
            country_gen_cats[code][gen] = {}
        country_gen_cats[code][gen][cat] = \
            country_gen_cats[code][gen].get(cat, 0) + 1

    cat_lgb_out = {
        cat: {
            "gen1_count": len(v["gen1"]),
            "gen2_count": len(v["gen2"]),
            "gen3_count": len(v["gen3"]),
            "mentions":   v["mentions"],
        }
        for cat, v in cat_lgb_sets.items()
    }

    cat_gen_aa_out = {
        gen: {
            cat: {
                "active_countries":   len(s["active_c"]),
                "archived_countries": len(s["archived_c"]),
                "active_mentions":    s["active_m"],
                "archived_mentions":  s["archived_m"],
            }
            for cat, s in cats.items()
        }
        for gen, cats in cat_gen_aa_sets.items()
    }

    categories_latest = {
        cat: {
            "countries_count": len(v["gen1"] | v["gen2"] | v["gen3"]),
            "mentions":        v["mentions"],
        }
        for cat, v in cat_lgb_sets.items()
    }

    print(f"   ✓ Tab2 built: {len(cat_lgb_out)} categories")

    # ── EU member states ───────────────────────────────────────────────
    print("🔄  Adding EU member states…")

    EU_CODE = "EEU"
    EU_MEMBER_ISO3 = [
        "AUT","BEL","BGR","HRV","CYP","CZE","DNK","EST","FIN","FRA",
        "DEU","GRC","HUN","IRL","ITA","LVA","LTU","LUX","MLT","NLD",
        "POL","PRT","ROU","SVK","SVN","ESP","SWE"
    ]
    EU_NAMES = {
        "AUT":"Austria","BEL":"Belgium","BGR":"Bulgaria","HRV":"Croatia",
        "CYP":"Cyprus","CZE":"Czechia","DNK":"Denmark","EST":"Estonia",
        "FIN":"Finland","FRA":"France","DEU":"Germany","GRC":"Greece",
        "HUN":"Hungary","IRL":"Ireland","ITA":"Italy","LVA":"Latvia",
        "LTU":"Lithuania","LUX":"Luxembourg","MLT":"Malta","NLD":"Netherlands",
        "POL":"Poland","PRT":"Portugal","ROU":"Romania","SVK":"Slovakia",
        "SVN":"Slovenia","ESP":"Spain","SWE":"Sweden"
    }

    eu_doc_ids_local = {
        doc_id: info for doc_id, info in doc_id_info.items()
        if info["code"] == EU_CODE
    }

    eu_best = None; eu_best_p = 0
    for doc_id, info in eu_doc_ids_local.items():
        if info["status"] != "Active": continue
        p = GEN_PRIORITY[info["gen"]]
        if p > eu_best_p:
            eu_best_p = p; eu_best = (doc_id, info)

    if eu_best:
        latest_eu_doc_id = eu_best[0]
        latest_eu_gen    = eu_best[1]["gen"]
        latest_eu_has_t  = eu_best[1].get("has_t", False)
        latest_eu_has_ghg = latest_eu_doc_id in doc_has_ghg_transport

        eu_gen_cats_local    = {}
        eu_latest_cats_local = {}

        for row in mit_sheet.iter_rows(min_row=2, values_only=True):
            if not row[M_DOCID]: break
            doc_id  = row[M_DOCID]
            code    = str(row[M_CODE] or "").strip()
            version = row[M_VERSION]
            cat     = row[M_CATEGORY]
            if code != EU_CODE or not cat: continue
            if doc_id not in eu_doc_ids_local: continue
            cat = str(cat).strip()
            gen = get_gen(version)
            if not gen: continue
            if gen not in eu_gen_cats_local:
                eu_gen_cats_local[gen] = {}
            eu_gen_cats_local[gen][cat] = eu_gen_cats_local[gen].get(cat, 0) + 1
            if doc_id == latest_eu_doc_id:
                eu_latest_cats_local[cat] = eu_latest_cats_local.get(cat, 0) + 1

        eu_gens_info = {}
        for row in doc_sheet.iter_rows(min_row=2, values_only=True):
            if not row[D_DOCID]: break
            if row[D_CODE] != EU_CODE or row[D_TYPE] != "NDC": continue
            gen = get_gen(row[D_VERSION])
            if not gen: continue
            has_t = str(row[D_TRANSPORT] or "").strip().lower() == "yes"
            ver   = str(row[D_VERSION]).strip()
            if gen not in eu_gens_info:
                eu_gens_info[gen] = {"has_transport": has_t, "version": ver}
            elif has_t:
                eu_gens_info[gen]["has_transport"] = True
                eu_gens_info[gen]["version"] = ver

        for iso3 in EU_MEMBER_ISO3:
            countries_tab1[iso3] = {
                "iso3":                    iso3,
                "name":                    EU_NAMES.get(iso3, iso3),
                "region":                  "Europe",
                "latest_active_gen":       latest_eu_gen,
                "latest_has_transport":    latest_eu_has_t,
                "latest_has_ghg_transport": latest_eu_has_ghg,
                "generations":             eu_gens_info,
                "covered_by_eu":           True,
            }
            country_gen_cats[iso3]    = eu_gen_cats_local.copy()
            country_latest_cats[iso3] = eu_latest_cats_local.copy()
            latest_gen_map[iso3]      = latest_eu_gen

        print(f"   ✓ Added {len(EU_MEMBER_ISO3)} EU member states "
              f"(latest: {latest_eu_gen}, transport: {latest_eu_has_t}, GHG: {latest_eu_has_ghg})")

    # ── Assemble output ────────────────────────────────────────────────
    output = {
        "metadata": {
            "total_possible_ndcs": 169,
            "last_updated":        str(date.today()),
            "data_source":         "GIZ-SLOCAT Transport Tracker Database",
            "gen_counts":          gen_counts,
        },
        "tab1": {
            "generations": tab1_generations,
            "countries":   countries_tab1,
        },
        "tab2": {
            "categories_latest":        categories_latest,
            "cat_latest_gen_breakdown": cat_lgb_out,
            "cat_gen_active_archived":  cat_gen_aa_out,
            "country_gen_cats":         country_gen_cats,
            "country_latest_cats":      country_latest_cats,
            "latest_gen_map":           latest_gen_map,
        },
    }

    wb.close()
    return output


# ============================================================================
# Entry point
# ============================================================================

def main():
    print("=" * 70)
    print("🚀  NDC Transport Tracker — Data Update")
    print("=" * 70)

    data_dir    = Path("data")
    excel_files = list(data_dir.glob("*.xlsx"))

    if not excel_files:
        print("❌  No .xlsx file found in data/")
        return 1

    excel_path = excel_files[0]
    print(f"\n📂  Excel file: {excel_path.name}")

    try:
        data = process_excel(excel_path)
    except Exception as exc:
        print(f"\n❌  Processing error: {exc}")
        raise

    output_dir = data_dir / "processed"
    output_dir.mkdir(exist_ok=True)

    json_path = output_dir / "data.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    print(f"\n💾  data.json saved ({json_path.stat().st_size:,} bytes)")
    print("\n" + "=" * 70)
    print("✅  Done!")
    print(f"   Countries:  {len(data['tab1']['countries'])}")
    print(f"   Categories: {len(data['tab2']['categories_latest'])}")
    gen = data['tab1']['generations']
    for g in ['gen1','gen2','gen3']:
        print(f"   {g}: {gen[g]['total_submitted']} submitted, "
              f"{gen[g]['with_transport']} transport, "
              f"{gen[g]['with_ghg_transport']} GHG transport")
    print("=" * 70)
    return 0


if __name__ == "__main__":
    sys.exit(main())
